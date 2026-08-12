import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.contratos import COLUMNS

# Larguras de coluna (em caracteres) por tipo de campo, para aproximar o autofit do Excel.
_COLUNAS_ESTREITAS = {
    "tipo_documento", "numero_contrato", "numero_contrato_sap", "numero_pedido_sap",
    "cnpj_contratante", "cnpj_contratada",
    "data_assinatura", "data_inicio_vigencia", "data_final_vigencia", "prazo_vigencia",
    "valor_total", "havera_reajuste",
}
_LARGURA_ESTREITA = 18
_LARGURA_LARGA = 45
_ALTURA_POR_LINHA = 12.5


def gerar_excel_formatado(df) -> bytes:
    """Gera um .xlsx seguindo a formatação exigida: Calibri 9, caixa alta, quebra de
    texto, alinhamento centralizado/à esquerda e cabeçalho verde em negrito."""
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Contratos"

    fonte_padrao = Font(name="Calibri", size=9)
    fonte_cabecalho = Font(name="Calibri", size=9, bold=True, color="000000")
    preenchimento_cabecalho = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    alinhamento = Alignment(horizontal="left", vertical="center", wrap_text=True)

    colunas = list(df.columns)
    ws.append([str(c).upper() for c in colunas])
    for cel in ws[1]:
        cel.font = fonte_cabecalho
        cel.fill = preenchimento_cabecalho
        cel.alignment = alinhamento

    chaves_por_indice = [chave for chave, _rotulo in COLUMNS]

    for _, linha in df.iterrows():
        valores = ["" if pd.isna(v) else str(v).upper() for v in linha]
        ws.append(valores)

    for idx, nome_coluna in enumerate(colunas, start=1):
        chave = chaves_por_indice[idx - 1] if idx - 1 < len(chaves_por_indice) else None
        largura = _LARGURA_ESTREITA if chave in _COLUNAS_ESTREITAS else _LARGURA_LARGA
        ws.column_dimensions[get_column_letter(idx)].width = largura

    for row_idx in range(2, ws.max_row + 1):
        max_linhas = 1
        for col_idx in range(1, ws.max_column + 1):
            celula = ws.cell(row=row_idx, column=col_idx)
            celula.font = fonte_padrao
            celula.alignment = alinhamento
            largura_col = ws.column_dimensions[get_column_letter(col_idx)].width or _LARGURA_LARGA
            texto = str(celula.value or "")
            linhas_texto = texto.count("\n") + 1
            linhas_estimadas = max(linhas_texto, -(-len(texto) // max(int(largura_col), 1)))
            max_linhas = max(max_linhas, linhas_estimadas)
        ws.row_dimensions[row_idx].height = max(15, max_linhas * _ALTURA_POR_LINHA)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
