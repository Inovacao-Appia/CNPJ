import json
import os
import re

import pandas as pd
import pdfplumber
import streamlit as st
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Cada prompt extrai um subconjunto de campos (pedir todos os campos de uma vez numa
# só chamada fazia o modelo confundir/misturar informações). Os 4 prompts abaixo
# espelham os documentos "PROMPT 1", "PROMPT 2.1", "PROMPT 2.2" e "PROMPT 3.1":
#   1 - Dados gerais (qualquer tipo de contrato)
#   2.1 - Vigências (Contrato de Obra + Ordem de Serviço)
#   2.2 - Vigências (Contrato de Prestação de Serviço, sem OS)
#   3.1 - Pagamentos: condição de pagamento, garantias/retenção e faturamento direto
# Todos rodam sobre o(s) mesmo(s) documento(s) e os resultados são combinados numa
# única linha, na ordem final de COLUMNS.
COLUMNS_1 = [
    ("tipo_documento", "TIPO DO DOCUMENTO (CONTRATO OU ADITIVO)"),
    ("numero_contrato", "NÚMERO DO CONTRATO"),
    ("numero_contrato_sap", "NÚMERO DO CONTRATO SAP (46)"),
    ("numero_pedido_sap", "NÚMERO DO PEDIDO DE COMPRA SAP (45)"),
    ("razao_social_contratante", "RAZÃO SOCIAL DA EMPRESA CONTRATANTE"),
    ("cnpj_contratante", "CNPJ DA EMPRESA CONTRATANTE"),
    ("razao_social_contratada", "RAZÃO SOCIAL DA EMPRESA CONTRATADA"),
    ("cnpj_contratada", "CNPJ DA EMPRESA CONTRATADA"),
    ("objeto_contrato", "OBJETO DO CONTRATO"),
    ("valor_total", "VALOR TOTAL DO CONTRATO - VALOR DA CONTRATAÇÃO"),
    ("detalhes_valor", "DETALHES DO VALOR DO CONTRATO"),
    ("valores_extras", "VALORES EXTRAS REEMBOLSÁVEIS"),
    ("havera_reajuste", "HAVERÁ REAJUSTE"),
    ("indice_reajuste", "ÍNDICE DE REAJUSTE"),
    ("local_servico", "LOCAL ONDE O SERVIÇO SERÁ REALIZADO"),
    ("gestor_contratante", "GESTOR DO CONTRATO POR PARTE DA CONTRATANTE (NOME E E-MAIL)"),
    ("gestor_contratada", "GESTOR DO CONTRATO POR PARTE DA CONTRATADA (NOME E E-MAIL)"),
    ("arquivo_origem", "ARQUIVO(S) DE ORIGEM"),
]

# Campos de vigência: mesmo conjunto de colunas para os contratos de Obra/OS (2.1) e
# de Prestação de Serviço (2.2) — o que muda entre eles é a REGRA de preenchimento
# (ver PROMPT_VIGENCIAS_OBRA_OS x PROMPT_VIGENCIAS_SERVICO), não os campos.
COLUMNS_2 = [
    ("tipo_documento", "TIPO DO DOCUMENTO (CONTRATO OU ADITIVO)"),
    ("numero_contrato", "NÚMERO DO CONTRATO"),
    ("data_assinatura", "DATA DA ASSINATURA DO DOCUMENTO"),
    ("data_inicio_vigencia", "DATA DE INÍCIO DA VIGÊNCIA"),
    ("data_final_vigencia", "DATA FINAL DA VIGÊNCIA"),
    ("prazo_vigencia", "PRAZO DE VIGÊNCIA"),
    ("periodo_medicao", "PERÍODO DE MEDIÇÃO MENSAL DO CONTRATO"),
    ("arquivo_origem", "ARQUIVO(S) DE ORIGEM"),
]

COLUMNS_3 = [
    ("tipo_documento", "TIPO DO DOCUMENTO (CONTRATO OU ADITIVO)"),
    ("numero_contrato", "NÚMERO DO CONTRATO"),
    ("condicao_pagamento", "CONDIÇÃO DE PAGAMENTO"),
    ("garantia_financeira", "GARANTIA FINANCEIRA / FIANÇA"),
    ("retencao_caucao", "RETENÇÃO CONTRATUAL / CAUÇÃO"),
    ("subcontratacao_faturamento_direto", "AUTORIZADA SUBCONTRATAÇÃO / FATURAMENTO DIRETO"),
    ("arquivo_origem", "ARQUIVO(S) DE ORIGEM"),
]

# Ordem final da planilha (rótulos vêm de COLUMNS_1/2/3 — fonte única de verdade).
_ORDEM_FINAL = [
    "tipo_documento",
    "numero_contrato",
    "numero_contrato_sap",
    "numero_pedido_sap",
    "razao_social_contratante",
    "cnpj_contratante",
    "razao_social_contratada",
    "cnpj_contratada",
    "objeto_contrato",
    "data_assinatura",
    "data_inicio_vigencia",
    "data_final_vigencia",
    "prazo_vigencia",
    "valor_total",
    "detalhes_valor",
    "valores_extras",
    "condicao_pagamento",
    "havera_reajuste",
    "indice_reajuste",
    "garantia_financeira",
    "retencao_caucao",
    "periodo_medicao",
    "local_servico",
    "subcontratacao_faturamento_direto",
    "gestor_contratante",
    "gestor_contratada",
    "arquivo_origem",
]
_rotulos_por_chave = dict(COLUMNS_1 + COLUMNS_2 + COLUMNS_3)
COLUMNS = [(chave, _rotulos_por_chave[chave]) for chave in _ORDEM_FINAL]
del _rotulos_por_chave

NAO_LOCALIZADO = "NÃO LOCALIZADO"

_JSON_KEYS_1 = ", ".join(f'"{chave}"' for chave, _ in COLUMNS_1)
_JSON_KEYS_2 = ", ".join(f'"{chave}"' for chave, _ in COLUMNS_2)
_JSON_KEYS_3 = ", ".join(f'"{chave}"' for chave, _ in COLUMNS_3)

# Contexto/identificação padrão: documento é exclusivamente CONTRATO ou ADITIVO
# (usado pelos prompts 1 - dados gerais, 2.2 - vigências de serviço e 3.1 - pagamentos).
_CONTEXTO_PADRAO = f"""CONTEXTO DO PROCESSAMENTO
- Cada execução deste prompt processa UM ÚNICO documento de cada vez.
- Todo documento fornecido será, exclusivamente, um CONTRATO ou um ADITIVO DE CONTRATO.
- Um ADITIVO deve ser tratado como uma linha totalmente separada e independente, com seus próprios campos extraídos exclusivamente a partir do conteúdo do próprio aditivo.
- NUNCA misture ou copie informações de outro documento (contrato principal, outro aditivo, OS, etc.) que porventura esteja anexo ao mesmo processamento. Extraia apenas do documento que está sendo analisado na presente execução.

IDENTIFICAÇÃO DO TIPO DE DOCUMENTO
- Identifique se o documento é um CONTRATO ou um ADITIVO e preencha o campo "tipo_documento" com "CONTRATO" ou "ADITIVO".
- Se for um ADITIVO, o campo "numero_contrato" deve conter o número do contrato principal seguido da identificação do aditivo, no formato "NÚMERO DO CONTRATO – ADITIVO Nº XX" (ex.: "ENG 778/2023 – ADITIVO Nº 01"). Se o aditivo não tiver número, use "ADITIVO" sem numeração.
- Se for um CONTRATO, o campo "numero_contrato" deve conter apenas o número do contrato."""

# Contexto/identificação para contratos de Obra: o documento pode ser CONTRATO, ADITIVO
# ou ORDEM DE SERVIÇO (usado apenas pelo prompt 2.1 - vigências de obra/OS).
_CONTEXTO_OBRA_OS = f"""CONTEXTO DO PROCESSAMENTO
- Todo documento fornecido será, exclusivamente, um CONTRATO, um ADITIVO DE CONTRATO ou uma ORDEM DE SERVIÇO (OS).
- Um ADITIVO deve ser tratado como uma linha totalmente separada e independente, com seus próprios campos extraídos exclusivamente a partir do conteúdo do próprio aditivo.

IDENTIFICAÇÃO DO TIPO DE DOCUMENTO
- Identifique se o documento é um CONTRATO ou um ADITIVO e preencha o campo "tipo_documento" com "CONTRATO" ou "ADITIVO".
- Se for um ADITIVO, o campo "numero_contrato" deve conter o número do contrato principal seguido da identificação do aditivo, no formato "NÚMERO DO CONTRATO – ADITIVO Nº XX" (ex.: "ENG 778/2023 – ADITIVO Nº 01"). Se o aditivo não tiver número, use "ADITIVO" sem numeração.
- Se for um CONTRATO, o campo "numero_contrato" deve conter apenas o número do contrato."""

# Bloco "REGRAS GERAIS" idêntico nos 4 prompts.
_REGRAS_GERAIS = f"""REGRAS GERAIS
1. Analise integralmente o documento enviado para esta execução.
2. Considere, além do contrato/aditivo principal, exclusivamente os anexos, DocuSign e demais arquivos que compõem ESTE documento (não de outros documentos processados em separado).
3. Para cada informação extraída, inclua no próprio texto do campo a origem do dado: cláusula (número e título, quando houver), anexo (quando aplicável) e a página do documento onde a informação foi localizada. Os documentos estão marcados no texto recebido com "=== {{NOME DO DOCUMENTO}} — PÁGINA {{N}} ==="; use esse nome de documento e número de página nas citações.
4. Quando uma informação estiver distribuída em mais de uma cláusula ou documento (do presente documento), cite todas as referências utilizadas.
5. Não faça suposições nem preencha informações por dedução. Extraia apenas dados expressamente previstos no documento, ou realize cálculos apenas quando houver regra específica para isso.
6. Caso uma informação não exista ou não seja localizada, preencha o campo com "{NAO_LOCALIZADO}".
7. Mantenha a redação objetiva, preservando o significado da cláusula contratual.
8. TODO o texto de todos os campos deve estar em CAIXA ALTA (maiúsculas), incluindo e-mails."""

# ---------------------------------------------------------------------------
# PROMPT 1 — Dados gerais (qualquer tipo de contrato)
# ---------------------------------------------------------------------------
PROMPT_DADOS_GERAIS = f"""Você atuará como um analista de contratos especializado na extração de informações contratuais.

{_CONTEXTO_PADRAO}

Sua função é analisar o documento fornecido para identificar, interpretar e extrair apenas as informações relacionadas aos campos listados abaixo.

{_REGRAS_GERAIS}

REGRAS ESPECÍFICAS POR CAMPO
MANTER A ORDEM DAS COLUNAS, CONFORME CAMPOS INDICADOS ABAIXO.

Observação: para os campos abaixo, NUNCA analisar documento que contém o nome "ORDEM DE SERVIÇO" ou "OS".

- Tipo do documento: informe "CONTRATO" ou "ADITIVO", conforme identificado.
- Número do contrato: se CONTRATO, informe o número do contrato. Se ADITIVO, informe "NÚMERO DO CONTRATO – ADITIVO Nº XX" (ou "– ADITIVO" se não houver numeração), citando cláusula/página onde a identificação foi localizada.
- Número do contrato SAP (46): informar o número de contrato SAP. Observação: número do contrato SAP inicia-se com 46, exemplo: 4600002220; alguns contratos podem não conter essa numeração, caso não encontrado preencher com "{NAO_LOCALIZADO}". Não registrar nesse campo quando iniciar com 45 (é o número do pedido de compra).
- Número do pedido de compra SAP (45): informar o número de pedido de compra SAP. Observação: número do pedido de compra SAP inicia-se com 45, exemplo: 4500052042; alguns contratos podem não conter essa numeração, caso não encontrado preencher com "{NAO_LOCALIZADO}". Não registrar nesse campo quando iniciar com 46 (é o número do contrato SAP).
- Razão Social da empresa Contratante.
- CNPJ da empresa Contratante.
- Razão Social da empresa Contratada.
- CNPJ da empresa Contratada.
- Objeto do contrato: identifique a cláusula que descreve o objeto contratual. Informe de forma fiel ao documento, resumindo apenas quando o texto for muito extenso, sem alterar o sentido. No caso de aditivo, descreva o objeto do aditivo (ex.: alteração de valor, prorrogação de vigência, inclusão de cláusula, etc.).
- Valor total do contrato - valor da contratação: informe o valor total, citando cláusula/página. Em aditivos que alteram valor, informe o novo valor total (e indique que se trata de saldo decorrente de aditivo quando aplicável).
- Detalhes do valor do contrato: informe valor total, valor mensal, valor unitário por serviço, adiantamentos e demais valores previstos. Em aditivos que alteram valor, informe o novo valor total e indique que se trata de saldo decorrente de aditivo quando aplicável.
- Valores extras reembolsáveis: informe valores adicionais previstos (reembolso de hospedagem, alimentação, viagens, diárias, deslocamentos e demais despesas reembolsáveis). Caso não exista previsão, informe "Não previstos".
- Haverá reajuste: informe "SIM" ou "NÃO", citando cláusula/página.
- Índice de reajuste: quando houver previsão de reajuste, informe o índice, a periodicidade e a cláusula/página. Quando não houver, repita a regra 5.
- Local onde o serviço será realizado: informe exatamente o local indicado no documento, sem interpretações.
- Gestor do contrato por parte da Contratante (nome e e-mail): identifique o profissional indicado pela CONTRATANTE como gestor do contrato, com respectivo e-mail. Quando houver mais de um, cite todos.
- Gestor do contrato por parte da Contratada (nome e e-mail): identifique o profissional indicado pela CONTRATADA como gestor/responsável técnico, com respectivo e-mail. Quando houver mais de um, cite todos.
- Arquivo(s) de origem: liste o(s) nome(s) do(s) arquivo(s) efetivamente utilizado(s) para extrair as informações deste documento (ex.: "ENG-778.23 - Afirma Engenharia e Projetos LTDA_contrato.pdf"). Se mais de um arquivo for fornecido, liste todos separados por ponto e vírgula.

IMPORTANTE
- Se o documento for um aditivo que trata APENAS de prorrogação de vigência, datas e demais campos podem permanecer com o mesmo conteúdo; ainda assim, extraia APENAS do próprio aditivo (transcreva o que o aditivo reproduzir/referenciar). Se o aditivo não reproduzir/especificar o campo, preencha com "{NAO_LOCALIZADO}" (não copie do contrato principal).

Responda EXCLUSIVAMENTE com um objeto JSON contendo exatamente estas chaves: {_JSON_KEYS_1}.
"""

# ---------------------------------------------------------------------------
# PROMPT 2.1 — Vigências (Contrato de Obra + Ordem de Serviço)
# ---------------------------------------------------------------------------
PROMPT_VIGENCIAS_OBRA_OS = f"""Você atuará como um analista de contratos especializado na extração de informações contratuais.

{_CONTEXTO_OBRA_OS}

Sua função é analisar o documento fornecido para identificar, interpretar e extrair apenas as informações relacionadas aos campos listados abaixo.

{_REGRAS_GERAIS}

REGRAS ESPECÍFICAS POR CAMPO
MANTER A ORDEM DAS COLUNAS, NA ORDEM DOS CAMPOS LISTADOS ABAIXO.

- Tipo do documento: informe "CONTRATO" ou "ADITIVO", conforme identificado.
- Número do contrato: se CONTRATO, informe o número do contrato. Se ADITIVO, informe "NÚMERO DO CONTRATO – ADITIVO Nº XX" (ou "– ADITIVO" se não houver numeração), citando cláusula/página onde a identificação foi localizada.
- Data da assinatura do documento: se houver informações de DocuSign, use a Data de Conclusão (Completed Date) do DOCUMENTO em análise (contrato ou aditivo). Caso não exista DocuSign, localize a página com as assinaturas e use a data ali constante. Analisar apenas o documento que contém o nome "CONTRATO" ou "ADITIVO" para esse campo. NUNCA analisar o documento que contém o nome "ORDEM DE SERVIÇO" ou "OS" para esse campo.
- Data de início da vigência: se o documento estabelecer que a vigência se inicia na data de assinatura da Ordem de Serviço (OS), localize a data de assinatura no documento da OS (Completed Date do DocuSign da OS, ou a data da página de assinaturas da OS caso não haja DocuSign) e utilize-a. Caso o documento informe uma data de início específica, utilize a data expressamente indicada. NUNCA analisar o documento que contém o nome "ADITIVO" para esse campo. Quando houver, analisar em conjunto os dois arquivos que contêm os nomes "Contrato" e "Ordem de Serviço" ou "OS", para preencher apenas este campo.
- Data final da vigência: identifique o prazo de vigência previsto (ex.: 12 meses, 365 dias, 24 meses). Se a vigência terminar após esse prazo contado da data de assinatura da OS, calcule a data final somando o prazo à data de assinatura da OS. Caso o documento informe uma data final específica, utilize-a diretamente. No documento que contém o nome "ADITIVO", se tratar apenas de prorrogação do prazo, trazer apenas a data final informada no aditivo.
- Prazo de vigência: informe o prazo previsto (dias, meses ou anos), citando cláusula, anexo e página. Em aditivos que alteram prazos, informe o novo prazo estabelecido pelo aditivo. Em aditivos de prorrogação, calcule a nova data final conforme o novo prazo estabelecido pelo aditivo.
- Período de medição mensal do contrato: informe a data inicial/final do período, periodicidade, data limite para apresentação da medição e cláusula/página. Caso não haja previsão, informe "Período de medição não previsto." Analisar apenas o documento que contém o nome "CONTRATO" ou "ADITIVO" para esse campo. NUNCA analisar o documento que contém o nome "ORDEM DE SERVIÇO" ou "OS" para esse campo.
- Arquivo(s) de origem: liste o(s) nome(s) do(s) arquivo(s) efetivamente utilizado(s) para extrair as informações deste documento. Se mais de um arquivo for fornecido, liste todos separados por ponto e vírgula.

Responda EXCLUSIVAMENTE com um objeto JSON contendo exatamente estas chaves: {_JSON_KEYS_2}.
"""

# ---------------------------------------------------------------------------
# PROMPT 2.2 — Vigências (Contrato de Prestação de Serviço, sem OS)
# ---------------------------------------------------------------------------
PROMPT_VIGENCIAS_SERVICO = f"""Você atuará como um analista de contratos especializado na extração de informações contratuais.

{_CONTEXTO_PADRAO}

Sua função é analisar o documento fornecido para identificar, interpretar e extrair apenas as informações relacionadas aos campos listados abaixo.

{_REGRAS_GERAIS}

REGRAS ESPECÍFICAS POR CAMPO
MANTER A ORDEM DAS COLUNAS, NA ORDEM DOS CAMPOS LISTADOS ABAIXO.

- Tipo do documento: informe "CONTRATO" ou "ADITIVO", conforme identificado.
- Número do contrato: se CONTRATO, informe o número do contrato. Se ADITIVO, informe "NÚMERO DO CONTRATO – ADITIVO Nº XX" (ou "– ADITIVO" se não houver numeração), citando cláusula/página onde a identificação foi localizada.
- Data da assinatura do documento: identificar e extrair a data de assinatura do documento em análise (contrato ou aditivo). Verifique se o documento contém informações de assinatura eletrônica do DocuSign. Caso exista DocuSign, localize o campo "Data de Conclusão (Completed Date)" referente ao documento em análise e extraia essa data como "Data da assinatura do documento". Caso o documento não contenha informações do DocuSign, localize a página de assinaturas do documento em análise. Extraia a data de assinatura constante na página de assinaturas. Caso existam várias datas de assinatura, utilize a data mais recente, que representa a conclusão da assinatura pelas partes.
- Data de início da vigência: identificar e extrair a data de início da vigência do contrato. Localize no contrato a data que indique explicitamente o início da vigência, utilizando expressões como "início da vigência", "vigência inicia em", "vigência terá início em", "o contrato entra em vigor em", "a partir de", "início de vigência" ou outras expressões equivalentes. Caso exista uma data de início da vigência expressamente informada, extraia essa data. Caso o contrato não informe explicitamente a data de início da vigência, mas estabeleça que a vigência se inicia na data de assinatura do contrato (ou utilize expressões equivalentes, como "na data de sua assinatura", "a partir da assinatura", "contados da assinatura"), utilize a data existente no campo "Data da assinatura do documento" como valor para este campo. Caso não seja possível identificar a data de início da vigência nem haja indicação de que ela corresponde à data de assinatura, informar "{NAO_LOCALIZADO}". NUNCA analisar o documento que contém o nome "ADITIVO" para esse campo.
- Data final da vigência: identifique o prazo de vigência previsto (ex.: 12 meses, 365 dias, 24 meses). Se a vigência terminar após esse prazo contado da data de assinatura do contrato, calcule a data final somando o prazo à data de assinatura do contrato. Caso o documento informe uma data final específica, utilize-a diretamente. Em aditivos de prorrogação, calcule a nova data final conforme o novo prazo estabelecido pelo aditivo. No documento que contém o nome "ADITIVO", se tratar apenas de prorrogação do prazo, trazer apenas a data final informada no aditivo.
- Prazo de vigência: informe o prazo previsto (dias, meses ou anos), citando cláusula, anexo e página. Em aditivos que alteram prazos, informe o novo prazo estabelecido pelo aditivo. Em aditivos de prorrogação, calcule a nova data final conforme o novo prazo estabelecido pelo aditivo.
- Período de medição mensal do contrato: informe a data inicial/final do período, periodicidade, data limite para apresentação da medição e cláusula/página. Caso não haja previsão, informe "Período de medição não previsto."
- Arquivo(s) de origem: liste o(s) nome(s) do(s) arquivo(s) efetivamente utilizado(s) para extrair as informações deste documento. Se mais de um arquivo for fornecido, liste todos separados por ponto e vírgula.

Responda EXCLUSIVAMENTE com um objeto JSON contendo exatamente estas chaves: {_JSON_KEYS_2}.
"""

# ---------------------------------------------------------------------------
# PROMPT 3.1 — Pagamentos: condição de pagamento, garantias/retenção e
# faturamento direto (todos os contratos)
# ---------------------------------------------------------------------------
PROMPT_PAGAMENTOS = f"""Você atuará como um analista de contratos especializado na extração de informações contratuais.

{_CONTEXTO_PADRAO}

Sua função é analisar o documento fornecido para identificar, interpretar e extrair apenas as informações relacionadas aos campos listados abaixo.

{_REGRAS_GERAIS}

REGRAS ESPECÍFICAS POR CAMPO
MANTER A ORDEM DAS COLUNAS, NA ORDEM DOS CAMPOS LISTADOS ABAIXO.

Observação: para os campos abaixo, NUNCA analisar documento que contém o nome "ORDEM DE SERVIÇO" ou "OS".

- Tipo do documento: informe "CONTRATO" ou "ADITIVO", conforme identificado.
- Número do contrato: se CONTRATO, informe o número do contrato. Se ADITIVO, informe "NÚMERO DO CONTRATO – ADITIVO Nº XX" (ou "– ADITIVO" se não houver numeração), citando cláusula/página onde a identificação foi localizada.
- Condição de pagamento: analise o contrato e demais documentos relacionados. Identifique todas as condições de pagamento previstas: se existe pagamento antecipado e o respectivo percentual ou valor; se o pagamento é realizado em parcela única ou em parcelas; se houver pagamento por etapas, marcos ou entregas do serviço, descreva cada etapa e o respectivo percentual ou valor a ser pago; se houver prazo para pagamento (ex.: 30 dias após emissão da nota fiscal, 15 dias após aceite da entrega etc.), informe o prazo. Caso não exista pagamento antecipado ou parcelamento por etapas, informe a condição de pagamento exatamente como descrita no contrato.
- Garantia Financeira/ Fiança: informe se há exigência de instrumento de garantia complementar (fiança bancária, seguro-garantia, apólice, nota promissória, garantia de adiantamento, garantia de performance, garantia em dinheiro ou outra modalidade). Preencha de forma objetiva, no padrão: (i) existência: "SIM", "NÃO" ou "NÃO APLICÁVEL"; (ii) modalidade(s) aceita(s); (iii) percentual e/ou valor da cobertura sobre o valor do contrato; (iv) momento de apresentação (ex.: na assinatura do contrato); (v) prazo de vigência e/ou restituição; (vi) cláusula/anexo/página de origem. A expressão "retenção" não determina sozinha esta coluna: prevalece o CONTEXTO ESTRUTURAL da cláusula.
- Retenção Contratual/ Caução: informe se há retenção mensal sobre as medições ou caução. Preencha de forma objetiva, no padrão: (i) existência: "SIM" ou "NÃO"; (ii) percentual da retenção sobre o valor bruto de cada medição e/ou valor fixo (ex.: R$ 5.000,00 ou 1%, prevalecendo o maior); (iii) base de cálculo (valor da medição, cada medição, última medição, etc.); (iv) prazo e condição de devolução (ex.: 180 dias após a medição final; 12 meses após o Termo de Recebimento Definitivo; sem correção monetária); (v) se a retenção pode ou não ser substituída por fiança bancária ou outra modalidade; (vi) cláusula/anexo/página de origem.
- Autorizada subcontratação / Faturamento direto: analise o documento em busca de cláusulas relacionadas à subcontratação, terceirização, cessão de atividades, faturamento direto ou pagamento direto a terceiros. Identifique se a subcontratação é: permitida sem restrições; permitida mediante autorização prévia da CONTRATANTE; permitida parcialmente (apenas para determinadas atividades); ou proibida. Identifique se existe previsão de faturamento direto, pagamento direto a subcontratados ou qualquer outra forma de faturamento por terceiros. Resuma as condições previstas, informando se a subcontratação é autorizada e em quais condições, e se há previsão de faturamento direto e quais são os requisitos para sua realização.
- Arquivo(s) de origem: liste o(s) nome(s) do(s) arquivo(s) efetivamente utilizado(s) para extrair as informações deste documento. Se mais de um arquivo for fornecido, liste todos separados por ponto e vírgula.

Responda EXCLUSIVAMENTE com um objeto JSON contendo exatamente estas chaves: {_JSON_KEYS_3}.
"""


def _get_client():
    api_key = st.secrets.get("OPENAI_API_KEY") or os.getenv("OPENAI_API_KEY")
    if not api_key:
        st.error("Chave OPENAI_API_KEY não configurada. Adicione em Settings → Secrets no Streamlit Cloud.")
        st.stop()
    return OpenAI(api_key=api_key)


# "OS" só conta como Ordem de Serviço quando é um token isolado (limites que não sejam
# letra/dígito dos dois lados) — evita falso positivo em nomes como "COSTA" ou "OSASCO".
_PADRAO_ORDEM_SERVICO = re.compile(r"ORDEM\s+DE\s+SERVI[CÇ]O|(?<![A-Z0-9])OS(?![A-Z0-9])", re.IGNORECASE)


def _e_ordem_servico(rotulo: str) -> bool:
    return bool(_PADRAO_ORDEM_SERVICO.search(rotulo))


def extrair_texto_documentos(arquivos: list[tuple[str, str]]) -> str:
    """Recebe lista de (rótulo_do_documento, caminho_pdf) e retorna o texto combinado,
    com marcadores de documento e página para permitir citação correta da origem."""
    partes = []
    for rotulo, caminho in arquivos:
        with pdfplumber.open(caminho) as pdf:
            for i, pagina in enumerate(pdf.pages, start=1):
                texto = pagina.extract_text()
                if texto:
                    partes.append(f"=== {rotulo.upper()} — PÁGINA {i} ===\n{texto}")
    return "\n\n".join(partes)


def _extrair_campos(client, prompt: str, texto: str, colunas: list[tuple[str, str]]) -> dict:
    response = client.chat.completions.create(
        model="gpt-4o",
        response_format={"type": "json_object"},
        temperature=0,
        messages=[
            {"role": "system", "content": prompt},
            {"role": "user", "content": f"Texto extraído do documento:\n{texto}"},
        ],
    )
    dados = json.loads(response.choices[0].message.content)

    resultado = {}
    for chave, _ in colunas:
        valor = dados.get(chave)
        resultado[chave] = str(valor).upper() if valor not in (None, "") else NAO_LOCALIZADO
    return resultado


def analisar_contrato(documentos: list[tuple[str, str]]) -> dict:
    """Roda os 4 prompts (cada um cobrindo um subconjunto de campos) sobre os
    documentos do contrato e combina os resultados numa única linha, na ordem de
    COLUMNS.

    - Dados gerais e Pagamentos (prompts 1 e 3.1) NUNCA devem analisar Ordem de
      Serviço, então recebem apenas o texto dos documentos que não são OS.
    - Vigências usa o prompt 2.1 (Obra + OS) quando há Ordem de Serviço no grupo,
      ou o prompt 2.2 (Prestação de Serviço) quando não há.
    """
    client = _get_client()

    documentos_sem_os = [(rotulo, caminho) for rotulo, caminho in documentos if not _e_ordem_servico(rotulo)]
    tem_os = len(documentos_sem_os) < len(documentos)

    texto_sem_os = extrair_texto_documentos(documentos_sem_os)
    texto_vigencias = extrair_texto_documentos(documentos) if tem_os else texto_sem_os
    prompt_vigencias = PROMPT_VIGENCIAS_OBRA_OS if tem_os else PROMPT_VIGENCIAS_SERVICO

    dados_gerais = _extrair_campos(client, PROMPT_DADOS_GERAIS, texto_sem_os, COLUMNS_1)
    dados_vigencias = _extrair_campos(client, prompt_vigencias, texto_vigencias, COLUMNS_2)
    dados_pagamentos = _extrair_campos(client, PROMPT_PAGAMENTOS, texto_sem_os, COLUMNS_3)

    combinado = {**dados_vigencias, **dados_pagamentos, **dados_gerais}
    return {chave: combinado.get(chave, NAO_LOCALIZADO) for chave, _ in COLUMNS}


def agrupar_arquivos_zip(tmpdir: str) -> list[tuple[str, list[tuple[str, str]]]]:
    """Varre o diretório extraído do ZIP e agrupa os PDFs por contrato:
    - cada subpasta (em qualquer profundidade) vira um grupo, contendo todos os PDFs
      encontrados recursivamente dentro dela;
    - cada PDF solto diretamente na raiz do ZIP vira seu próprio grupo (documento único).
    """
    grupos: list[tuple[str, list[tuple[str, str]]]] = []

    itens_raiz = sorted(os.listdir(tmpdir))
    for item in itens_raiz:
        caminho_item = os.path.join(tmpdir, item)

        if os.path.isdir(caminho_item):
            documentos = []
            for root, _, files in os.walk(caminho_item):
                for file in sorted(files):
                    if file.lower().endswith(".pdf"):
                        rotulo = os.path.splitext(file)[0]
                        documentos.append((rotulo, os.path.join(root, file)))
            if documentos:
                grupos.append((item, documentos))
        elif item.lower().endswith(".pdf"):
            rotulo = os.path.splitext(item)[0]
            grupos.append((rotulo, [(rotulo, caminho_item)]))

    return grupos


# Larguras de coluna (em caracteres) por tipo de campo, para aproximar o autofit do Excel.
_COLUNAS_ESTREITAS = {
    "tipo_documento", "numero_contrato", "numero_contrato_sap", "numero_pedido_sap",
    "cnpj_contratante", "cnpj_contratada",
    "data_assinatura", "data_inicio_vigencia", "data_final_vigencia", "prazo_vigencia",
    "valor_total", "havera_reajuste",
}
_LARGURA_ESTREITA = 18
_LARGURA_LARGA = 45
_CARACTERES_POR_LINHA = 60
_ALTURA_POR_LINHA = 12.5


def gerar_excel_formatado(df) -> bytes:
    """Gera um .xlsx seguindo a formatação exigida: Calibri 9, caixa alta, quebra de
    texto, alinhamento centralizado/à esquerda e cabeçalho verde em negrito."""
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Contratos"

    fonte_padrao = Font(name="Calibri", size=9)
    fonte_cabecalho = Font(name="Calibri", size=9, bold=True, color="000000")
    preenchimento_cabecalho = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    alinhamento = Alignment(horizontal="left", vertical="center", wrap_text=True)

    colunas = list(df.columns)
    ws.append([str(c).upper() for c in colunas])
    for cel in ws[1]:
        cel.font = fonte_cabecalho
        cel.fill = preenchimento_cabecalho
        cel.alignment = alinhamento

    chaves_por_indice = [chave for chave, _rotulo in COLUMNS]

    for _, linha in df.iterrows():
        valores = ["" if pd.isna(v) else str(v).upper() for v in linha]
        ws.append(valores)

    for idx, nome_coluna in enumerate(colunas, start=1):
        chave = chaves_por_indice[idx - 1] if idx - 1 < len(chaves_por_indice) else None
        largura = _LARGURA_ESTREITA if chave in _COLUNAS_ESTREITAS else _LARGURA_LARGA
        ws.column_dimensions[get_column_letter(idx)].width = largura

    for row_idx in range(2, ws.max_row + 1):
        max_linhas = 1
        for col_idx in range(1, ws.max_column + 1):
            celula = ws.cell(row=row_idx, column=col_idx)
            celula.font = fonte_padrao
            celula.alignment = alinhamento
            largura_col = ws.column_dimensions[get_column_letter(col_idx)].width or _LARGURA_LARGA
            texto = str(celula.value or "")
            linhas_texto = texto.count("\n") + 1
            linhas_estimadas = max(linhas_texto, -(-len(texto) // max(int(largura_col), 1)))
            max_linhas = max(max_linhas, linhas_estimadas)
        ws.row_dimensions[row_idx].height = max(15, max_linhas * _ALTURA_POR_LINHA)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
